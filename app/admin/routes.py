import os
import io
from flask import render_template, redirect, url_for, flash, request, current_app, send_file, jsonify
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from app import db
from app.admin import bp
from app.admin.forms import UserForm, TeamForm, SpeciesForm, SkillForm, TrainingPathForm, ImportForm, AddUserToTeamForm, TrainingValidationForm, AttendeeValidationForm, CompetencyValidationForm, RoleForm, ContinuousTrainingEventForm, BatchValidateUserContinuousTrainingForm, ValidateUserContinuousTrainingEntryForm # Added new forms
from app.training.forms import TrainingSessionForm # Import TrainingSessionForm
from app.models import User, Team, Species, Skill, TrainingPath, TrainingPathSkill, ExternalTraining, TrainingRequest, TrainingRequestStatus, ExternalTrainingStatus, Competency, TrainingSession, SkillPracticeEvent, Complexity, ExternalTrainingSkillClaim, TrainingSessionTutorSkill, tutor_skill_association, Permission, Role, ContinuousTrainingEvent, ContinuousTrainingEventStatus, UserContinuousTraining, UserContinuousTrainingStatus # Added new models and enums
from app.decorators import permission_required
from sqlalchemy import func, extract, case
import openpyxl # Import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from datetime import datetime, timedelta, timezone # Import datetime, timedelta
from collections import defaultdict
import json # Import json
import re
from app.email import send_email # Import send_email

@bp.route('/continuous_training_events')
@login_required
@permission_required('continuous_training_manage')
def manage_continuous_training_events():
    status_filter = request.args.get('status', '', type=str)
    
    query = db.session.query(
        ContinuousTrainingEvent,
        func.count(case((UserContinuousTraining.status == UserContinuousTrainingStatus.APPROVED, UserContinuousTraining.id), else_=None)).label('approved_attendees_count')
    ).outerjoin(UserContinuousTraining, ContinuousTrainingEvent.id == UserContinuousTraining.event_id)

    if status_filter:
        query = query.filter(ContinuousTrainingEvent.status == ContinuousTrainingEventStatus[status_filter])

    events_with_attendee_count = query.group_by(ContinuousTrainingEvent.id).order_by(ContinuousTrainingEvent.event_date.desc()).all()
    
    statuses = [s.name for s in ContinuousTrainingEventStatus]
    
    return render_template('admin/manage_continuous_training_events.html', 
                           title='Gérer les Événements de Formation Continue', 
                           events=events_with_attendee_count, # Pass events with attendee count
                           statuses=statuses,
                           current_status=status_filter)

@bp.route('/continuous_training_events/<int:event_id>/attendees')
@login_required
@permission_required('continuous_training_manage')
def get_continuous_training_event_attendees(event_id):
    event = ContinuousTrainingEvent.query.get_or_404(event_id)
    attendees = []
    for user_ct in event.user_attendances.filter_by(status=UserContinuousTrainingStatus.APPROVED).all():
        attendees.append({
            'id': user_ct.user.id,
            'full_name': user_ct.user.full_name,
            'email': user_ct.user.email,
            'validated_hours': user_ct.validated_hours
        })
    return jsonify(attendees)

@bp.route('/continuous_training_events/validate_quick/<int:event_id>', methods=['POST'])
@login_required
@permission_required('continuous_training_manage') # Assuming this permission allows validation
def validate_quick_continuous_training_event(event_id):
    event = ContinuousTrainingEvent.query.get_or_404(event_id)

    # Check if all required info is present
    # Required fields: title, training_type, event_date, duration_hours
    if event.title and event.training_type and event.event_date and event.duration_hours is not None: # Check for None for duration_hours
        event.status = ContinuousTrainingEventStatus.APPROVED
        event.validator = current_user
        db.session.add(event)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Event approved successfully!'})
    else:
        return jsonify({'success': False, 'message': 'Missing information, please edit the event.', 'redirect_to_edit': url_for('admin.edit_continuous_training_event', event_id=event.id)})

@bp.route('/continuous_training_events/<int:event_id>/remove_attendee/<int:user_id>', methods=['POST'])
@login_required
@permission_required('continuous_training_manage') # Assuming this permission allows managing attendees
def remove_continuous_training_attendee(event_id, user_id):
    user_ct = UserContinuousTraining.query.filter_by(event_id=event_id, user_id=user_id).first_or_404()
    db.session.delete(user_ct)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Attendee removed successfully!'})

@bp.route('/continuous_training_events/add', methods=['GET', 'POST'])
@login_required
@permission_required('continuous_training_manage')
def add_continuous_training_event():
    form = ContinuousTrainingEventForm()
    if form.validate_on_submit():
        attachment_path = None
        if form.attachment.data:
            filename = secure_filename(f"ct_event_{datetime.utcnow().timestamp()}_{form.attachment.data.filename}")
            upload_path = os.path.join(current_app.root_path, 'static', 'uploads', 'continuous_training_events')
            os.makedirs(upload_path, exist_ok=True)
            form.attachment.data.save(os.path.join(upload_path, filename))
            attachment_path = f"uploads/continuous_training_events/{filename}"

        event = ContinuousTrainingEvent(
            title=form.title.data,
            description=form.description.data,
            training_type=form.training_type.data,
            location=form.location.data,
            event_date=form.event_date.data,
            duration_hours=form.duration_hours.data,
            attachment_path=attachment_path,
            creator=current_user,
            status=ContinuousTrainingEventStatus.PENDING # Events are pending until validated by an admin/validator
        )
        db.session.add(event)
        db.session.commit()
        flash('Événement de formation continue ajouté avec succès !', 'success')
        return redirect(url_for('admin.manage_continuous_training_events'))
    return render_template('admin/continuous_training_event_form.html', title='Ajouter un Événement de Formation Continue', form=form)

@bp.route('/continuous_training_events/edit/<int:event_id>', methods=['GET', 'POST'])
@login_required
@permission_required('continuous_training_manage')
def edit_continuous_training_event(event_id):
    event = ContinuousTrainingEvent.query.get_or_404(event_id)
    form = ContinuousTrainingEventForm(obj=event)
    if form.validate_on_submit():
        # Handle "Validate Event" button submission
        if request.form.get('validate_event') == 'true':
            if not form.duration_hours.data:
                flash('La durée en heures est obligatoire pour valider l\'événement.', 'danger')
                return render_template('admin/continuous_training_event_form.html', title='Éditer un Événement de Formation Continue', form=form, event=event)
            
            event.status = ContinuousTrainingEventStatus.APPROVED
            event.validator = current_user
            flash('Événement de formation continue validé avec succès !', 'success')
        else: # Regular "Save Event" submission
            flash('Événement de formation continue mis à jour avec succès !', 'success')

        event.title = form.title.data
        event.description = form.description.data
        event.training_type = form.training_type.data
        event.location = form.location.data
        event.event_date = form.event_date.data
        event.duration_hours = form.duration_hours.data

        if form.attachment.data:
            # Delete old attachment if exists
            if event.attachment_path:
                old_path = os.path.join(current_app.root_path, 'static', event.attachment_path)
                if os.path.exists(old_path):
                    os.remove(old_path)

            filename = secure_filename(f"ct_event_{datetime.utcnow().timestamp()}_{form.attachment.data.filename}")
            upload_path = os.path.join(current_app.root_path, 'static', 'uploads', 'continuous_training_events')
            os.makedirs(upload_path, exist_ok=True)
            form.attachment.data.save(os.path.join(upload_path, filename))
            event.attachment_path = f"uploads/continuous_training_events/{filename}"
        
        db.session.commit()
        return redirect(url_for('admin.manage_continuous_training_events'))
    elif request.method == 'GET':
        form.title.data = event.title
        form.description.data = event.description
        form.training_type.data = event.training_type.name
        form.location.data = event.location
        form.event_date.data = event.event_date
        form.duration_hours.data = event.duration_hours

    return render_template('admin/continuous_training_event_form.html', title='Éditer un Événement de Formation Continue', form=form, event=event)

@bp.route('/continuous_training_events/delete/<int:event_id>', methods=['POST'])
@login_required
@permission_required('continuous_training_manage')
def delete_continuous_training_event(event_id):
    event = ContinuousTrainingEvent.query.get_or_404(event_id)
    db.session.delete(event)
    db.session.commit()
    flash('Événement de formation continue supprimé avec succès !', 'success')
    return redirect(url_for('admin.manage_continuous_training_events'))

@bp.route('/initial_regulatory_trainings')
@login_required
@permission_required('initial_regulatory_training_manage')
def manage_initial_regulatory_trainings():
    initial_trainings = InitialRegulatoryTraining.query.order_by(InitialRegulatoryTraining.training_date.desc()).all()
    return render_template('admin/manage_initial_regulatory_trainings.html', title='Gérer les Formations Réglementaires Initiales', initial_trainings=initial_trainings)

@bp.route('/initial_regulatory_trainings/add', methods=['GET', 'POST'])
@login_required
@permission_required('initial_regulatory_training_manage')
def add_initial_regulatory_training():
    form = AdminInitialRegulatoryTrainingForm()
    if form.validate_on_submit():
        attachment_path = None
        if form.attachment.data:
            filename = secure_filename(f"{form.user.data.id}_initial_reg_training_{datetime.utcnow().timestamp()}_{form.attachment.data.filename}")
            upload_path = os.path.join(current_app.root_path, 'static', 'uploads', 'initial_regulatory_training')
            os.makedirs(upload_path, exist_ok=True)
            form.attachment.data.save(os.path.join(upload_path, filename))
            attachment_path = f"uploads/initial_regulatory_training/{filename}"

        initial_training = InitialRegulatoryTraining(
            user=form.user.data,
            level=InitialRegulatoryTrainingLevel[form.level.data],
            training_date=form.training_date.data,
            attachment_path=attachment_path
        )
        db.session.add(initial_training)
        db.session.commit()
        flash('Formation réglementaire initiale ajoutée avec succès !', 'success')
        return redirect(url_for('admin.manage_initial_regulatory_trainings'))
    return render_template('admin/initial_regulatory_training_form.html', title='Ajouter une Formation Réglementaire Initiale', form=form)

@bp.route('/initial_regulatory_trainings/edit/<int:training_id>', methods=['GET', 'POST'])
@login_required
@permission_required('initial_regulatory_training_manage')
def edit_initial_regulatory_training(training_id):
    initial_training = InitialRegulatoryTraining.query.get_or_404(training_id)
    form = AdminInitialRegulatoryTrainingForm(obj=initial_training)
    if form.validate_on_submit():
        initial_training.user = form.user.data
        initial_training.level = InitialRegulatoryTrainingLevel[form.level.data]
        initial_training.training_date = form.training_date.data

        if form.attachment.data:
            # Delete old attachment if exists
            if initial_training.attachment_path:
                old_path = os.path.join(current_app.root_path, 'static', initial_training.attachment_path)
                if os.path.exists(old_path):
                    os.remove(old_path)

            filename = secure_filename(f"{form.user.data.id}_initial_reg_training_{datetime.utcnow().timestamp()}_{form.attachment.data.filename}")
            upload_path = os.path.join(current_app.root_path, 'static', 'uploads', 'initial_regulatory_training')
            os.makedirs(upload_path, exist_ok=True)
            form.attachment.data.save(os.path.join(upload_path, filename))
            initial_training.attachment_path = f"uploads/initial_regulatory_training/{filename}"
        
        db.session.commit()
        flash('Formation réglementaire initiale mise à jour avec succès !', 'success')
        return redirect(url_for('admin.manage_initial_regulatory_trainings'))
    elif request.method == 'GET':
        form.user.data = initial_training.user
        form.level.data = initial_training.level.name
        form.training_date.data = initial_training.training_date

    return render_template('admin/initial_regulatory_training_form.html', title='Éditer une Formation Réglementaire Initiale', form=form, initial_training=initial_training)

@bp.route('/initial_regulatory_trainings/delete/<int:training_id>', methods=['POST'])
@login_required
@permission_required('initial_regulatory_training_manage')
def delete_initial_regulatory_training(training_id):
    initial_training = InitialRegulatoryTraining.query.get_or_404(training_id)
    db.session.delete(initial_training)
    db.session.commit()
    flash('Formation réglementaire initiale supprimée avec succès !', 'success')
    return redirect(url_for('admin.manage_initial_regulatory_trainings'))

@bp.route('/validate_continuous_trainings')
@login_required
@permission_required('continuous_training_validate')
def validate_continuous_trainings():
    pending_user_cts = UserContinuousTraining.query.filter_by(status=UserContinuousTrainingStatus.PENDING).order_by(UserContinuousTraining.validation_date.desc()).all()
    form = BatchValidateUserContinuousTrainingForm()
    
    # Populate form for GET request
    for user_ct in pending_user_cts:
        entry_form = ValidateUserContinuousTrainingEntryForm()
        entry_form.user_ct_id = user_ct.id
        entry_form.user_full_name = user_ct.user.full_name
        entry_form.event_title = user_ct.event.title
        entry_form.event_date = user_ct.event.event_date.strftime('%Y-%m-%d %H:%M')
        entry_form.attendance_attachment_path = user_ct.attendance_attachment_path
        entry_form.validated_hours = user_ct.event.duration_hours # Pre-fill with event's duration
        entry_form.status = UserContinuousTrainingStatus.PENDING.name
        form.entries.append_entry(entry_form)

    return render_template('admin/validate_continuous_trainings.html', title='Valider Formations Continues', form=form, pending_user_cts=pending_user_cts)

@bp.route('/validate_continuous_trainings/batch', methods=['POST'])
@login_required
@permission_required('continuous_training_validate')
def batch_validate_continuous_trainings():
    form = BatchValidateUserContinuousTrainingForm()
    if form.validate_on_submit():
        for entry in form.entries:
            user_ct = UserContinuousTraining.query.get(entry.user_ct_id.data)
            if user_ct:
                user_ct.validated_hours = entry.validated_hours.data
                user_ct.status = UserContinuousTrainingStatus[entry.status.data]
                user_ct.validated_by = current_user
                user_ct.validation_date = datetime.utcnow()
                db.session.add(user_ct)
        db.session.commit()
        flash('Formations continues validées avec succès !', 'success')
        return redirect(url_for('admin.validate_continuous_trainings'))
    flash('Erreur lors de la validation des formations continues.', 'danger')
    return redirect(url_for('admin.validate_continuous_trainings'))

@bp.route('/validate_continuous_trainings/single/<int:user_ct_id>', methods=['POST'])
@login_required
@permission_required('continuous_training_validate')
def single_validate_continuous_training(user_ct_id):
    user_ct = UserContinuousTraining.query.get_or_404(user_ct_id)
    
    data = request.get_json()
    if not data:
        return jsonify({'success': False, 'message': 'No JSON data provided.'}), 400

    validated_hours_str = data.get('validated_hours')
    status_str = data.get('status')

    # Manual validation and type conversion for validated_hours
    try:
        validated_hours = float(validated_hours_str)
        if validated_hours < 0:
            raise ValueError("Validated hours cannot be negative.")
    except (ValueError, TypeError):
        return jsonify({'success': False, 'message': 'Heures validées invalides. Doit être un nombre positif.'}), 400
    
    try:
        status = UserContinuousTrainingStatus[status_str]
    except KeyError:
        return jsonify({'success': False, 'message': 'Statut invalide.'}), 400

    user_ct.validated_hours = validated_hours
    user_ct.status = status
    user_ct.validated_by = current_user
    user_ct.validation_date = datetime.utcnow()
    
    try:
        db.session.add(user_ct)
        db.session.commit()
        
        current_app.logger.debug(f"UserContinuousTraining {user_ct_id} validated successfully.")
        return jsonify({'success': True, 'message': 'Formation continue validée avec succès !'})
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Database error during single validation for user_ct_id {user_ct_id}: {e}")
        return jsonify({'success': False, 'message': f'Erreur de base de données lors de la validation: {str(e)}'}), 500

@bp.route('/validate_continuous_trainings/reject/<int:user_ct_id>', methods=['POST'])
@login_required
@permission_required('continuous_training_validate')
def reject_continuous_training(user_ct_id):
    print(f"Rejecting user_ct_id: {user_ct_id}")
    user_ct = UserContinuousTraining.query.get_or_404(user_ct_id)
    print(f"UserContinuousTraining found: {user_ct}")
    user_ct.status = UserContinuousTrainingStatus.REJECTED
    user_ct.validated_by = current_user
    user_ct.validation_date = datetime.utcnow()
    db.session.add(user_ct)
    db.session.commit()
    print(f"UserContinuousTraining {user_ct_id} status updated to REJECTED")
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        print("Returning JSON response")
        return jsonify({'success': True, 'message': 'Continuous training rejected successfully!'})
    print("Returning redirect response")
    flash('Formation continue rejetée avec succès !', 'info')
    return redirect(url_for('admin.validate_continuous_trainings'))

@bp.route('/')
@bp.route('/index')
@login_required
@permission_required('admin_access')
def index():
    # Metrics for the cards
    pending_requests_count = TrainingRequest.query.filter_by(status=TrainingRequestStatus.PENDING).count()
    pending_external_trainings_count = ExternalTraining.query.filter_by(status=ExternalTrainingStatus.PENDING).count()
    skills_without_tutors_count = Skill.query.filter(~Skill.tutors.any()).count()
    proposed_skills_count = TrainingRequest.query.filter_by(status=TrainingRequestStatus.PROPOSED_SKILL).count()
    pending_user_approvals_count = User.query.filter_by(is_approved=False).count()
    pending_continuous_training_validations_count = UserContinuousTraining.query.filter_by(status=UserContinuousTrainingStatus.PENDING).count()
    pending_continuous_event_requests_count = ContinuousTrainingEvent.query.filter_by(status=ContinuousTrainingEventStatus.PENDING).count()

    # Placeholder for more complex metrics
    recycling_needed_count = 0
    users_needing_recycling_set = set() # Keep this to pass to the template if needed
    
    recycling_map = defaultdict(set)
    for comp in Competency.query.options(db.joinedload(Competency.skill)).all():
        if comp.needs_recycling:
            recycling_map[comp.user_id].add(comp.skill_id)
            recycling_needed_count += 1
            users_needing_recycling_set.add(comp.user)

    users_needing_recycling = list(users_needing_recycling_set)
    
    sessions_to_be_finalized_count = TrainingSession.query.filter(
        TrainingSession.start_time < datetime.now(timezone.utc),
        TrainingSession.status != 'Realized'
    ).count()

    # Logic for sessions this month (now next session)
    now = datetime.now(timezone.utc)
    next_session = TrainingSession.query.filter(TrainingSession.start_time > now).order_by(TrainingSession.start_time.asc()).first()

    # Data for the tables
    users = User.query.options(db.joinedload(User.teams), db.joinedload(User.teams_as_lead)).order_by(User.full_name).all() # Eagerly load teams and teams_as_lead
    skills = Skill.query.options(db.joinedload(Skill.species)).order_by(Skill.name).all()
    pending_training_requests = TrainingRequest.query.options(
        db.joinedload(TrainingRequest.requester),
        db.joinedload(TrainingRequest.skills_requested),
        db.joinedload(TrainingRequest.species_requested)
    )
    teams = Team.query.all()

    all_continuous_events = (db.session.query(
        ContinuousTrainingEvent,
        func.count(case((UserContinuousTraining.status == UserContinuousTrainingStatus.APPROVED, UserContinuousTraining.id), else_=None)).label('approved_attendees_count')
    )
    .outerjoin(UserContinuousTraining, ContinuousTrainingEvent.id == UserContinuousTraining.event_id)
    .group_by(ContinuousTrainingEvent.id)
    .order_by(ContinuousTrainingEvent.event_date.desc())
    .all())

    # Data for the validation table
    validation_form = BatchValidateUserContinuousTrainingForm()
    pending_user_cts = UserContinuousTraining.query.filter_by(status=UserContinuousTrainingStatus.PENDING).all()
    
    for user_ct in pending_user_cts:
        entry_form = ValidateUserContinuousTrainingEntryForm()
        entry_form.user_ct_id = user_ct.id
        entry_form.user_full_name = user_ct.user.full_name
        entry_form.event_title = user_ct.event.title
        entry_form.event_date = user_ct.event.event_date.strftime('%Y-%m-%d')
        entry_form.attendance_attachment_path = user_ct.attendance_attachment_path
        entry_form.validated_hours = user_ct.event.duration_hours
        entry_form.status = user_ct.status.name # Set the current status
        validation_form.entries.append_entry(entry_form)


    return render_template('admin/admin_dashboard.html',
                           title='Admin Dashboard',
                           pending_requests_count=pending_requests_count,
                           pending_external_trainings_count=pending_external_trainings_count,
                           skills_without_tutors_count=skills_without_tutors_count,
                           proposed_skills_count=proposed_skills_count,
                           pending_user_approvals_count=pending_user_approvals_count,
                           pending_continuous_training_validations_count=pending_continuous_training_validations_count,
                           pending_continuous_event_requests_count=pending_continuous_event_requests_count,
                           recycling_needed_count=recycling_needed_count,
                           sessions_to_be_finalized_count=sessions_to_be_finalized_count,
                           next_session=next_session,
                           users=users,
                           skills=skills,
                           pending_training_requests=pending_training_requests,
                           users_needing_recycling=users_needing_recycling,
                           teams=teams,
                           recycling_map=recycling_map,
                           all_continuous_events=all_continuous_events,
                           validation_form=validation_form,
                           pending_user_cts=pending_user_cts)


@bp.route('/pending_users')
@login_required
@permission_required('user_manage')
def pending_users():
    pending_users = User.query.filter_by(is_approved=False).all()
    return render_template('admin/pending_users.html', title='Pending User Approvals', pending_users=pending_users)

@bp.route('/approve_user/<int:user_id>', methods=['POST'])
@login_required
@permission_required('user_manage')
def approve_user(user_id):
    user = User.query.get_or_404(user_id)
    user.is_approved = True
    user_role = Role.query.filter_by(name='User').first()
    if user_role and user_role not in user.roles:
        user.roles.append(user_role)
    db.session.commit()
    flash(f'User {user.full_name} approved successfully!', 'success')

    # Send approval email to user
    send_email('[Training Manager] Account Approved',
               sender=current_app.config['MAIL_USERNAME'],
               recipients=[user.email],
               text_body=render_template('email/registration_approved.txt', user=user)),
    return redirect(url_for('admin.edit_user', id=user_id))

@bp.route('/reject_user/<int:user_id>', methods=['POST'])
@login_required
@permission_required('user_manage')
def reject_user(user_id):
    user = User.query.get_or_404(user_id)
    db.session.delete(user) # Delete the user if rejected
    db.session.commit()
    flash(f'User {user.full_name} rejected and deleted.', 'info')

    # Send rejection email to user
    send_email('[Training Manager] Account Rejected',
               sender=current_app.config['MAIL_USERNAME'],
               recipients=[user.email],
               text_body=render_template('email/registration_rejected.txt', user=user),
               html_body=render_template('email/registration_rejected.html', user=user))

    return redirect(url_for('admin.edit_user', id=user_id))

@bp.route('/teams')
@login_required
@permission_required('team_manage')
def manage_teams():
    teams = Team.query.all()
    return render_template('admin/manage_teams.html', title='Manage Teams', teams=teams)


@bp.route('/users/add', methods=['GET', 'POST'])
@login_required
@permission_required('user_manage')
def add_user():
    form = UserForm()
    if form.validate_on_submit():
        user = User(full_name=form.full_name.data, email=form.email.data,
                    is_admin=form.is_admin.data, study_level=form.study_level.data)
        user.set_password(form.password.data)
        
        # Generate API key for new user
        user.generate_api_key() # Add this line
        
        # Handle many-to-many relationships
        user.teams = form.teams.data
        user.teams_as_lead = form.teams_as_lead.data
        user.roles = form.roles.data # Assign roles

        db.session.add(user)
        db.session.commit()

        # Handle assigned training paths
        user.assigned_training_paths = form.assigned_training_paths.data

        # Process training requests for each selected training path
        for training_path in form.assigned_training_paths.data:
            for tps in training_path.skills_association:
                if tps.skill: # Ensure skill exists
                    training_request = TrainingRequest(
                        requester=user,
                        status=TrainingRequestStatus.PENDING,
                        notes=f"Automatically generated from Training Path: {training_path.name}"
                    )
                    db.session.add(training_request)
                    db.session.flush() # Assign an ID to the new training_request
                    training_request.skills_requested.append(tps.skill)
                    training_request.species_requested.append(training_path.species)
                else:
                    current_app.logger.warning(f"TrainingPathSkill {tps.training_path_id}-{tps.skill_id} has no associated skill. Skipping.")
        db.session.commit()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True, 
                'message': 'User added successfully!',
                'user': {
                    'id': user.id,
                    'full_name': user.full_name,
                    'email': user.email,
                    'is_admin': user.is_admin,
                    'study_level': user.study_level,
                    'teams': [t.name for t in user.teams],
                    'teams_as_lead': [lt.name for lt in user.teams_as_lead],
                    'roles': [r.name for r in user.roles]
                }
            })

        flash('User added successfully!', 'success')
        return redirect(url_for('admin.manage_users'))
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_template('admin/_user_form_fields.html', form=form)

    return render_template('admin/user_form.html', title='Add User', form=form)

@bp.route('/users/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('user_manage')
def edit_user(id):
    user = User.query.get_or_404(id)
    form = UserForm(original_email=user.email)
    if form.validate_on_submit():
        user.full_name = form.full_name.data
        user.email = form.email.data
        if form.password.data:
            user.set_password(form.password.data)
        user.is_admin = form.is_admin.data
        user.study_level = form.study_level.data
        
        # Generate API key if missing
        if user.api_key is None: # Add this check
            user.generate_api_key() # Add this line
            
        # Get currently assigned training paths before updating
        current_training_paths = set(user.assigned_training_paths)

        # Handle many-to-many relationships
        user.teams = form.teams.data
        user.teams_as_lead = form.teams_as_lead.data
        user.assigned_training_paths = form.assigned_training_paths.data
        user.roles = form.roles.data # Assign roles

        # Identify newly added training paths
        new_training_paths = set(form.assigned_training_paths.data)
        added_training_paths = new_training_paths - current_training_paths

        # Create training requests for newly added paths
        for training_path in added_training_paths:
            for tps in training_path.skills_association:
                if tps.skill: # Ensure skill exists
                    training_request = TrainingRequest(
                        requester=user,
                        status=TrainingRequestStatus.PENDING,
                        notes=f"Automatically generated from Training Path: {training_path.name}"
                    )
                    db.session.add(training_request)
                    db.session.flush() # Assign an ID to the new training_request
                    training_request.skills_requested.append(tps.skill)
                    training_request.species_requested.append(training_path.species)
                else:
                    current_app.logger.warning(f"TrainingPathSkill {tps.training_path_id}-{tps.skill_id} has no associated skill. Skipping.")

        db.session.commit()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True, 
                'message': 'User updated successfully!',
                'user': {
                    'id': user.id,
                    'full_name': user.full_name,
                    'email': user.email,
                    'is_admin': user.is_admin,
                    'study_level': user.study_level,
                    'teams': [t.name for t in user.teams],
                    'teams_as_lead': [lt.name for lt in user.teams_as_lead],
                    'roles': [r.name for r in user.roles]
                }
            })

        flash('User updated successfully!', 'success')
        return redirect(url_for('admin.manage_users'))
    elif request.method == 'GET':
        form.full_name.data = user.full_name
        form.email.data = user.email
        form.is_admin.data = user.is_admin
        form.study_level.data = user.study_level
        
        # Pre-populate many-to-many fields
        form.teams.data = user.teams
        form.teams_as_lead.data = user.teams_as_lead
        form.assigned_training_paths.data = user.assigned_training_paths
        form.roles.data = user.roles # Pre-populate roles
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_template('admin/_user_form_fields.html', form=form)
    return render_template('admin/user_form.html', title='Edit User', form=form)

@bp.route('/users/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('user_manage')
def delete_user(id):
    user = User.query.get_or_404(id)
    
    # Delete associated records with NOT NULL foreign keys
    TrainingRequest.query.filter_by(requester_id=user.id).delete()
    Competency.query.filter_by(user_id=user.id).delete()
    ExternalTraining.query.filter_by(user_id=user.id).delete()
    SkillPracticeEvent.query.filter_by(user_id=user.id).delete()

    # Clear many-to-many relationships
    user.teams.clear()
    user.teams_as_lead.clear()
    user.assigned_training_paths.clear()
    user.tutored_skills.clear()
    user.attended_training_sessions.clear()
    user.tutored_training_sessions.clear()

    db.session.delete(user)
    db.session.commit()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest': # Check if the request is an AJAX request
        return jsonify({'success': True, 'message': 'User deleted successfully!'})
    flash('User deleted successfully!', 'success')
    return redirect(url_for('admin.manage_users'))

# Role Management
@bp.route('/roles')
@login_required
@permission_required('role_manage')
def manage_roles():
    roles = Role.query.all()
    return render_template('admin/manage_roles.html', title='Manage Roles', roles=roles)

@bp.route('/roles/add', methods=['GET', 'POST'])
@login_required
@permission_required('role_manage')
def add_role():
    form = RoleForm()
    if form.validate_on_submit():
        role = Role(name=form.name.data, description=form.description.data)
        role.permissions = form.permissions.data
        db.session.add(role)
        db.session.commit()
        flash('Role added successfully!', 'success')
        return redirect(url_for('admin.manage_roles'))
    return render_template('admin/role_form.html', title='Add Role', form=form)

@bp.route('/roles/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('role_manage')
def edit_role(id):
    role = Role.query.get_or_404(id)
    form = RoleForm(original_name=role.name)
    if form.validate_on_submit():
        role.name = form.name.data
        role.description = form.description.data
        role.permissions = form.permissions.data
        db.session.commit()
        flash('Role updated successfully!', 'success')
        return redirect(url_for('admin.manage_roles'))
    elif request.method == 'GET':
        form.name.data = role.name
        form.description.data = role.description
        form.permissions.data = role.permissions
    return render_template('admin/role_form.html', title='Edit Role', form=form)

@bp.route('/roles/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('role_manage')
def delete_role(id):
    role = Role.query.get_or_404(id)
    db.session.delete(role)
    db.session.commit()
    flash('Role deleted successfully!', 'success')
    return redirect(url_for('admin.manage_roles'))

# Permission Management
@bp.route('/permissions')
@login_required
@permission_required('permission_manage')
def manage_permissions():
    permissions = Permission.query.all()
    return render_template('admin/manage_permissions.html', title='Manage Permissions', permissions=permissions)

# Team Management

@bp.route('/teams/add', methods=['GET', 'POST'])
@login_required
@permission_required('team_manage')
def add_team():
    form = TeamForm()
    if form.validate_on_submit():
        team = Team(name=form.name.data)
        db.session.add(team)
        db.session.flush() # Flush to assign an ID to the new team

        # Handle many-to-many relationships
        team.members = form.members.data
        team.team_leads = form.team_leads.data

        db.session.commit()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True, 
                'message': 'Team added successfully!',
                'team': {
                    'id': team.id,
                    'name': team.name,
                    'members': [m.full_name for m in team.members],
                    'team_leads': [l.full_name for l in team.team_leads]
                }
            })

        flash('Team added successfully!', 'success')
        return redirect(url_for('admin.manage_teams'))
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_template('admin/_team_form_fields.html', form=form, api_key=current_user.api_key)

    return render_template('admin/team_form.html', title='Add Team', form=form)

@bp.route('/teams/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('team_manage')
def edit_team(id):
    team = Team.query.get_or_404(id)
    form = TeamForm(original_name=team.name)
    if form.validate_on_submit():
        team.name = form.name.data
        
        # Handle many-to-many relationships
        team.members = form.members.data
        team.team_leads = form.team_leads.data

        db.session.commit()
        flash('Team updated successfully!', 'success')
        return redirect(url_for('admin.index')) # Redirect to admin dashboard
    elif request.method == 'GET':
        form.name.data = team.name
        
        # Pre-populate many-to-many fields
        form.members.data = team.members
        form.team_leads.data = team.team_leads
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_template('admin/_team_form_fields.html', form=form, team=team, api_key=current_user.api_key)
    return render_template('admin/team_form.html', title='Edit Team', form=form)

@bp.route('/teams/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('team_manage')
def delete_team(id):
    team = Team.query.get_or_404(id)
    
    db.session.delete(team)
    db.session.commit()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'message': 'Team deleted successfully!'})

    flash('Team deleted successfully!', 'success')
    return redirect(url_for('admin.index')) # Redirect to admin dashboard
    flash('Team deleted successfully!', 'success')
    return redirect(url_for('admin.manage_teams'))

@bp.route('/team/<int:team_id>/add_users', methods=['GET', 'POST'])
@login_required
@permission_required('team_manage')
def add_users_to_team(team_id):
    team = Team.query.get_or_404(team_id)
    form = AddUserToTeamForm()

    if form.validate_on_submit():
        selected_users = form.users.data
        for user in selected_users:
            if user not in team.members: # Prevent adding duplicates
                team.members.append(user)
        db.session.commit()
        flash(f'{len(selected_users)} user(s) added to team {team.name} successfully!', 'success')
        return jsonify({'success': True, 'message': f'{len(selected_users)} user(s) added to team {team.name} successfully!'})

    # For GET request or form validation failure
    # Filter users to only show those not already in this team
    # This query needs to be updated to reflect the many-to-many relationship
    form.users.query = User.query.filter(~User.teams.any(id=team.id)).order_by(User.full_name).all()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_template('admin/_add_users_to_team_form.html', form=form, team=team)
    
    flash('Error adding users to team.', 'danger')
    return redirect(url_for('admin.index')) # Redirect to admin dashboard on error

# Species Management
@bp.route('/species')
@login_required
@permission_required('species_manage')
def manage_species():
    species_list = Species.query.all()
    return render_template('admin/manage_species.html', title='Manage Species', species_list=species_list)

@bp.route('/species/add', methods=['GET', 'POST'])
@login_required
@permission_required('species_manage')
def add_species():
    form = SpeciesForm()
    if form.validate_on_submit():
        species = Species(name=form.name.data)
        db.session.add(species)
        db.session.commit()
        flash('Species added successfully!', 'success')
        return redirect(url_for('admin.manage_species'))
    return render_template('admin/species_form.html', title='Add Species', form=form)

@bp.route('/species/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('species_manage')
def edit_species(id):
    species = Species.query.get_or_404(id)
    form = SpeciesForm(original_name=species.name)
    if form.validate_on_submit():
        species.name = form.name.data
        db.session.commit()
        flash('Species updated successfully!', 'success')
        return redirect(url_for('admin.manage_species'))
    elif request.method == 'GET':
        form.name.data = species.name
    return render_template('admin/species_form.html', title='Edit Species', form=form)

@bp.route('/species/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('species_manage')
def delete_species(id):
    species = Species.query.get_or_404(id)
    db.session.delete(species)
    db.session.commit()
    flash('Species deleted successfully!', 'success')
    return redirect(url_for('admin.manage_species'))

# Skill Management
@bp.route('/skills')
@login_required
@permission_required('skill_manage')
def manage_skills():
    skills_query = db.session.query(
        Skill,
        func.count(func.distinct(Competency.user_id)).label('user_count'),
        func.count(func.distinct(case((Competency.needs_recycling == True, Competency.user_id), else_=None))).label('recycling_count'),
        func.count(func.distinct(tutor_skill_association.c.user_id)).label('tutor_count')
    ).outerjoin(Competency, Skill.id == Competency.skill_id) \
     .outerjoin(tutor_skill_association, Skill.id == tutor_skill_association.c.skill_id) \
     .group_by(Skill.id)

    skill_name = request.args.get('skill_name', '')
    if skill_name:
        skills_query = skills_query.filter(Skill.name.ilike(f'%{skill_name}%'))

    needs_recycling = request.args.get('needs_recycling', 'false').lower() == 'true'
    if needs_recycling:
        skills_query = skills_query.having(func.count(func.distinct(case(((Competency.needs_recycling == True), Competency.user_id), else_=None))) > 0)

    skills_data = skills_query.order_by(Skill.name).all()
    form = ImportForm() # Instantiate the form

    return render_template('admin/import_export_skills.html', title='Manage Skills', skills_data=skills_data, needs_recycling=needs_recycling, skill_name=skill_name, form=form)

@bp.route('/api/skills')
@login_required
@permission_required('skill_manage')
def api_skills():
    search = request.args.get('q', '')
    if search:
        query = Skill.query.filter(Skill.name.ilike(f'%{search}%'))
    else:
        query = Skill.query
    skills = query.order_by(Skill.name).all()
    return jsonify([{'id': s.id, 'text': s.name} for s in skills])

@bp.route('/skills/add', methods=['GET', 'POST'])
@login_required
@permission_required('skill_manage')
def add_skill():
    form = SkillForm()
    proposal_id = request.args.get('from_proposal')
    proposal_to_delete = None
    if proposal_id:
        proposal_to_delete = TrainingRequest.query.get_or_404(proposal_id)
        if request.method == 'GET':
            # Regex to extract name and description
            match = re.match(r"Proposed Skill: (.*) - Description: (.*)", proposal_to_delete.notes)
            if match:
                form.name.data = match.group(1)
                form.description.data = match.group(2)

    if form.validate_on_submit():
        skill = Skill(name=form.name.data, description=form.description.data,
                      validity_period_months=form.validity_period_months.data,
                      complexity=form.complexity.data,
                      reference_urls_text=form.reference_urls_text.data,
                      training_videos_urls_text=form.training_videos_urls_text.data,
                      potential_external_tutors_text=form.potential_external_tutors_text.data)
        
        if form.protocol_attachment.data:
            filename = secure_filename(form.protocol_attachment.data.filename)
            upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'protocols')
            os.makedirs(upload_folder, exist_ok=True)
            file_path = os.path.join(upload_folder, filename)
            form.protocol_attachment.data.save(file_path)
            skill.protocol_attachment_path = os.path.join('uploads', 'protocols', filename)

        skill.species = form.species.data

        db.session.add(skill)
        
        # If the skill was created from a proposal, delete the proposal
        if proposal_to_delete:
            db.session.delete(proposal_to_delete)

        db.session.commit()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': 'Skill added successfully!',
                'skill': {
                    'id': skill.id,
                    'name': skill.name,
                    'description': skill.description
                }
            })

        flash('Skill added successfully!', 'success')
        return redirect(url_for('admin.manage_skills'))
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if form.errors:
            return jsonify({'success': False, 'form_html': render_template('admin/_skill_form_fields.html', form=form)})
        return render_template('admin/_skill_form_fields.html', form=form)

    return render_template('admin/skill_form.html', title='Add Skill', form=form)

@bp.route('/skills/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('skill_manage')
def edit_skill(id):
    skill = Skill.query.get_or_404(id)
    form = SkillForm(original_name=skill.name)
    


    if form.validate_on_submit():
        skill.name = form.name.data
        skill.description = form.description.data
        skill.validity_period_months = form.validity_period_months.data
        skill.complexity = form.complexity.data
        skill.reference_urls_text = form.reference_urls_text.data
        skill.training_videos_urls_text = form.training_videos_urls_text.data
        skill.potential_external_tutors_text = form.potential_external_tutors_text.data

        if form.protocol_attachment.data:
            filename = secure_filename(form.protocol_attachment.data.filename)
            upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'protocols')
            os.makedirs(upload_folder, exist_ok=True)
            file_path = os.path.join(upload_folder, filename)
            form.protocol_attachment.data.save(file_path)
            skill.protocol_attachment_path = os.path.join('uploads', 'protocols', filename)
        
        skill.species = form.species.data
        db.session.commit()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'message': 'Skill updated successfully!',
                'skill': {
                    'id': skill.id,
                    'name': skill.name,
                    'description': skill.description,
                    'species': [s.name for s in skill.species]
                }
            })

        flash('Skill updated successfully!', 'success')
        return redirect(url_for('admin.manage_skills'))
    
    elif request.method == 'GET':
        form.name.data = skill.name
        form.description.data = skill.description
        form.validity_period_months.data = skill.validity_period_months
        form.complexity.data = skill.complexity.name
        form.reference_urls_text.data = skill.reference_urls_text
        form.training_videos_urls_text.data = skill.training_videos_urls_text
        form.potential_external_tutors_text.data = skill.potential_external_tutors_text
        form.species.data = skill.species
        

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        if form.errors:
            return jsonify({'success': False, 'form_html': render_template('admin/_skill_form_fields.html', form=form)})
        return render_template('admin/_skill_form_fields.html', form=form)
    
    return render_template('admin/skill_form.html', title='Edit Skill', form=form)

@bp.route('/skills/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('skill_manage')
def delete_skill(id):
    skill = Skill.query.get_or_404(id)
    # Delete associated ExternalTrainingSkillClaim records first
    ExternalTrainingSkillClaim.query.filter_by(skill_id=skill.id).delete()
    db.session.delete(skill)
    db.session.commit()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'message': 'Skill deleted successfully!'})
    flash('Skill deleted successfully!', 'success')
    return redirect(url_for('admin.manage_skills'))

# Training Path Management
@bp.route('/skills/<int:skill_id>/users/<string:user_type>')
@login_required
@permission_required('skill_manage') # Assuming viewing skill users is part of skill management
def skill_users(skill_id, user_type):
    skill = Skill.query.get_or_404(skill_id)
    if user_type == 'competent':
        users = User.query.join(Competency, User.id == Competency.user_id).filter(Competency.skill_id == skill_id).all()
        title = f"Users with skill: {skill.name}"
    elif user_type == 'in_training':
        users = User.query.join(TrainingRequest, User.id == TrainingRequest.requester_id).filter(TrainingRequest.skills_requested.any(id=skill_id), TrainingRequest.status == TrainingRequestStatus.PENDING).all()
        title = f"Users in training for: {skill.name}"
    elif user_type == 'tutors':
        users = User.query.join(tutor_skill_association).filter(tutor_skill_association.c.skill_id == skill_id).all()
        title = f"Tutors for: {skill.name}"
    else:
        return redirect(url_for('admin.manage_skills'))
    return render_template('admin/skill_users.html', title=title, users=users, skill=skill)



@bp.route('/training_paths')
@login_required
@permission_required('training_path_manage')
def manage_training_paths():
    training_paths = TrainingPath.query.all()
    return render_template('admin/manage_training_paths.html', title='Manage Training Paths', training_paths=training_paths)

@bp.route('/training_paths/add', methods=['GET', 'POST'])
@login_required
@permission_required('training_path_manage')
def add_training_path():
    form = TrainingPathForm()
    if form.validate_on_submit():
        training_path = TrainingPath(name=form.name.data, description=form.description.data)
        training_path.species_id = form.species.data.id # Set species_id directly
        db.session.add(training_path)
        
        skills_data = json.loads(form.skills_json.data)
        
        # Server-side check for duplicate skills
        skill_ids_in_form = [skill_data['skill_id'] for skill_data in skills_data]
        if len(skill_ids_in_form) != len(set(skill_ids_in_form)):
            flash('Duplicate skills found in the training path. Please ensure each skill is unique.', 'danger')
            db.session.rollback()
            return redirect(url_for('admin.add_training_path'))

        for skill_data in skills_data:
            skill = Skill.query.get(skill_data['skill_id'])
            if skill:
                tps = TrainingPathSkill(
                    skill=skill,
                    order=skill_data['order']
                )
                training_path.skills_association.append(tps)

        db.session.add(training_path)
        db.session.commit()
        flash('Training Path added successfully!', 'success')
        return redirect(url_for('admin.manage_training_paths'))
        
    all_skills = Skill.query.order_by(Skill.name).all()
    all_species = Species.query.order_by(Species.name).all()
    all_skills_json = [{'id': s.id, 'name': s.name} for s in all_skills]
    
    # Ensure skills_json is an empty array for new forms
    if not form.skills_json.data:
        form.skills_json.data = '[]'

    return render_template('admin/training_path_form.html', title='Add Training Path', form=form, all_skills=all_skills_json, all_species=all_species, training_path=None)

@bp.route('/training_paths/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@permission_required('training_path_manage')
def edit_training_path(id):
    training_path = TrainingPath.query.get_or_404(id)
    form = TrainingPathForm(original_name=training_path.name)
    if form.validate_on_submit():
        training_path.name = form.name.data
        training_path.description = form.description.data
        training_path.species = form.species.data # Update species
        
        # Clear existing skills and re-add from form
        training_path.skills_association.clear()
        
        skills_data = json.loads(form.skills_json.data)

        # Server-side check for duplicate skills
        skill_ids_in_form = [skill_data['skill_id'] for skill_data in skills_data]
        if len(skill_ids_in_form) != len(set(skill_ids_in_form)):
            flash('Duplicate skills found in the training path. Please ensure each skill is unique.', 'danger')
            db.session.rollback()
            return redirect(url_for('admin.edit_training_path', id=training_path.id))

        for skill_data in skills_data:
            skill = Skill.query.get(skill_data['skill_id'])
            if skill:
                tps = TrainingPathSkill(
                    skill=skill,
                    order=skill_data['order']
                )
                training_path.skills_association.append(tps)

        db.session.commit()
        flash('Training Path updated successfully!', 'success')
        return redirect(url_for('admin.manage_training_paths'))
        
    elif request.method == 'GET':
        form.name.data = training_path.name
        form.description.data = training_path.description
        form.species.data = training_path.species # Pre-populate species
        
        skills_json_data = []
        for assoc in training_path.skills_association:
            skills_json_data.append({
                'skill_id': assoc.skill.id,
                'name': assoc.skill.name,
                'order': assoc.order
            })
        form.skills_json.data = json.dumps(skills_json_data)

    all_skills = Skill.query.order_by(Skill.name).all()
    all_species = Species.query.order_by(Species.name).all()
    all_skills_json = [{'id': s.id, 'name': s.name} for s in all_skills]
    return render_template('admin/training_path_form.html', title='Edit Training Path', form=form, all_skills=all_skills_json, all_species=all_species, training_path=training_path)

@bp.route('/training_paths/delete/<int:id>', methods=['POST'])
@login_required
@permission_required('training_path_manage')
def delete_training_path(id):
    training_path = TrainingPath.query.get_or_404(id)
    db.session.delete(training_path)
    db.session.commit()
    flash('Training Path deleted successfully!', 'success')
    return redirect(url_for('admin.manage_training_paths'))

# Import/Export Functionality (Placeholders)
@bp.route('/import_export_users', methods=['GET', 'POST'])
@login_required
@permission_required('user_manage')
def import_export_users():
    form = ImportForm()
    form.update_existing.label.text = "Update existing users if emails match?"
    if form.validate_on_submit():
        if form.import_file.data:
            file = form.import_file.data
            filename = secure_filename(file.filename)
            
            if filename.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                
                users_imported = 0
                users_updated = 0
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)): # Skip header
                    try:
                        # Assuming Excel columns: full_name, email, password, is_admin, is_team_lead, team_name
                        full_name, email, password, is_admin_str, is_team_lead_str, team_name = row
                        
                        user = User.query.filter_by(email=email).first()
                        if user is None:
                            is_admin = str(is_admin_str).lower() == 'true'
                            is_team_lead = str(is_team_lead_str).lower() == 'true'
                            
                            user = User(full_name=full_name, email=email, is_admin=is_admin)
                            user.set_password(password)
                            
                            if team_name:
                                team = Team.query.filter_by(name=team_name).first()
                                if not team:
                                    team = Team(name=team_name)
                                    db.session.add(team)
                                
                                user.teams.append(team)
                                if is_team_lead:
                                    user.teams_as_lead.append(team)
                            
                            db.session.add(user)
                            users_imported += 1
                        elif form.update_existing.data:
                            user.full_name = full_name
                            user.is_admin = str(is_admin_str).lower() == 'true'
                            if password:
                                user.set_password(password)
                            
                            is_team_lead = str(is_team_lead_str).lower() == 'true'
                            
                            # Clear existing teams and leadership roles
                            user.teams.clear()
                            user.teams_as_lead.clear()

                            if team_name:
                                team = Team.query.filter_by(name=team_name).first()
                                if not team:
                                    team = Team(name=team_name)
                                    db.session.add(team)
                                
                                user.teams.append(team)
                                if is_team_lead:
                                    user.teams_as_lead.append(team)
                            users_updated += 1

                    except Exception as e:
                        flash(f"Error importing row {row_idx+1}: {row} - {e}", 'danger')
                        db.session.rollback()
                        continue
                db.session.commit()
                flash(f"{users_imported} users imported, {users_updated} users updated successfully from Excel!", 'success')
            else:
                flash('Unsupported file format. Please upload an XLSX file.', 'danger')
            
            return redirect(url_for('admin.index')) # Redirect to admin dashboard after import

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_template('admin/_import_form.html', form=form, 
                               action_url=url_for('admin.import_export_users'),
                               template_url=url_for('admin.download_user_import_template_xlsx'))

    return render_template('admin/import_export_users.html', title='Import/Export Users', form=form)





@bp.route('/export_users_xlsx')
@login_required
@permission_required('user_manage')
def export_users_xlsx():
    users = User.query.all()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Users"

    # Write header
    sheet.append(['full_name', 'email', 'password', 'is_admin', 'is_team_lead', 'team_name'])

    # Write data
    for user in users:
        sheet.append([user.full_name, user.email, '', user.is_admin, bool(user.teams_as_lead), user.teams[0].name if user.teams else ''])
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'users_export_{datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")}.xlsx')

@bp.route('/download_user_import_template_xlsx')
@login_required
@permission_required('user_manage')
def download_user_import_template_xlsx():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "User Import Template"

    headers = ['full_name', 'email', 'password', 'is_admin', 'is_team_lead', 'team_name']
    sheet.append(headers)

    # Data validation for is_admin and is_team_lead columns
    dv_boolean = DataValidation(type="list", formula1='"TRUE,FALSE"', allow_blank=True)
    dv_boolean.add('D2:D1048576') # is_admin column
    dv_boolean.add('E2:E1048576') # is_team_lead column
    sheet.add_data_validation(dv_boolean)

    # Data validation for team_name (list of existing teams)
    team_names = [t.name for t in Team.query.all()]
    if team_names:
        dv_teams = DataValidation(type="list", formula1='"' + ','.join(team_names) + '"', allow_blank=True)
        dv_teams.add('F2:F1048576') # team_name column
        sheet.add_data_validation(dv_teams)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='user_import_template.xlsx')

@bp.route('/import_export_skills', methods=['GET', 'POST'])
@login_required
@permission_required('skill_manage')
def import_export_skills():
    form = ImportForm()
    if form.validate_on_submit():
        if form.import_file.data:
            file = form.import_file.data
            filename = secure_filename(file.filename)
            
            if filename.endswith('.xlsx'):
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                
                skills_imported = 0
                skills_updated = 0
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)): # Skip header
                    try:
                                # Assuming Excel columns: name, description, validity_period_months, complexity, reference_urls_text, training_videos_urls_text, potential_external_tutors_text, species_names
                        name, description, validity_period_months_str, complexity_str, reference_urls_text, training_videos_urls_text, potential_external_tutors_text, species_names_str = row                        
                        skill = Skill.query.filter_by(name=name).first()
                        
                        if skill is None:
                            validity_period_months = int(validity_period_months_str) if validity_period_months_str else None
                            complexity = Complexity(complexity_str) if complexity_str else Complexity.SIMPLE
                            
                            skill = Skill(
                                name=name,
                                description=description,
                                validity_period_months=validity_period_months,
                                complexity=complexity,
                                reference_urls_text=reference_urls_text,
                                training_videos_urls_text=training_videos_urls_text,
                                potential_external_tutors_text=potential_external_tutors_text
                            )
                            db.session.add(skill)
                            # The flush is needed to get the skill.id before adding species and tutors
                            db.session.flush()

                            if species_names_str:
                                species_names = [s.strip() for s in str(species_names_str).split(',')]
                                for species_name in species_names:
                                    species_obj = Species.query.filter_by(name=species_name).first()
                                    if species_obj:
                                        skill.species.append(species_obj)
                                    else:
                                        flash(f"Species '{species_name}' not found for skill '{name}'. It will be skipped.", 'warning')
                            skills_imported += 1
                        elif form.update_existing.data:
                            # Update existing skill
                            skill.description = description
                            skill.validity_period_months = int(validity_period_months_str) if validity_period_months_str else None
                            skill.complexity = Complexity(complexity_str) if complexity_str else Complexity.SIMPLE
                            skill.reference_urls_text = reference_urls_text
                            skill.training_videos_urls_text = training_videos_urls_text
                            skill.potential_external_tutors_text = potential_external_tutors_text

                            # Handle species (clear and re-add for updates)
                            skill.species.clear()
                            if species_names_str:
                                species_names = [s.strip() for s in str(species_names_str).split(',')]
                                for species_name in species_names:
                                    species_obj = Species.query.filter_by(name=species_name).first()
                                    if species_obj:
                                        skill.species.append(species_obj)
                                    else:
                                        flash(f"Species '{species_name}' not found for skill '{name}'. It will be skipped.", 'warning')
                            

                            skills_updated += 1
                        else:
                            flash(f"Skill '{name}' already exists and 'Update existing' was not checked. Skipping.", 'info')
                            continue

                    except Exception as e:
                        flash(f"Error importing row {row_idx+1}: {row} - {e}", 'danger')
                        db.session.rollback()
                        continue
                db.session.commit()
                flash(f"{skills_imported} skills imported and {skills_updated} skills updated successfully from Excel!", 'success')
            else:
                flash('Unsupported file format. Please upload an XLSX file.', 'danger')
            
            return redirect(url_for('admin.index')) # Redirect to admin dashboard after import
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render_template('admin/_import_form.html', form=form,
                               action_url=url_for('admin.import_export_skills'),
                               template_url=url_for('admin.download_skill_import_template'))

    return render_template('admin/import_export_skills.html', title='Import/Export Skills', form=form)
@bp.route('/download_skill_import_template')
@login_required
@permission_required('skill_manage')
def download_skill_import_template():
    # Get data for dropdowns
    complexity_values = [c.name for c in Complexity]
    species_names = [s.name for s in Species.query.all()]


    workbook = openpyxl.Workbook()
    sheet = workbook.active
    headers = [
        'name', 'description', 'validity_period_months', 'complexity',
        'reference_urls_text', 'training_videos_urls_text',
        'potential_external_tutors_text', 'species_names'
    ]
    sheet.append(headers)

    # Create data validation for 'complexity'
    dv_complexity = DataValidation(type="list", formula1='"' + ','.join(complexity_values) + '"', allow_blank=True)
    dv_complexity.add('D2:D1048576') # Apply to column D (Complexity) from row 2 onwards
    sheet.add_data_validation(dv_complexity)

    # Create data validation for 'species_names' (assuming comma-separated list)
    # This is a bit trickier for multi-select, so we'll provide a hint in the comment
    # For now, just a list of existing species for single selection or as a guide
    # if species_names:
    #     dv_species = DataValidation(type="list", formula1='"' + ','.join(species_names) + '"', allow_blank=True)
    #     dv_species.add('H2:H1048576') # Apply to column H (species_names) from row 2 onwards
    #     sheet.add_data_validation(dv_species)
    


    # Add a comment to guide users for multi-select fields
    sheet['H1'].comment = openpyxl.comments.Comment("For multiple species, separate names with commas (e.g., 'Species A, Species B')", "Admin")



    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='skill_import_template.xlsx')



@bp.route('/export_skills_xlsx')
@login_required
@permission_required('skill_manage')
def export_skills_xlsx():
    skills = Skill.query.all()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Skills"

    headers = [
        'name', 'description', 'validity_period_months', 'complexity',
        'reference_urls_text', 'training_videos_urls_text',
        'potential_external_tutors_text', 'species_names'
    ]
    sheet.append(headers)

    # Get data for dropdowns (same as import template)
    complexity_values = [c.name for c in Complexity]
    species_names_list = [s.name for s in Species.query.all()]


    # Create data validation for 'complexity'
    dv_complexity = DataValidation(type="list", formula1='"' + ','.join(complexity_values) + '"', allow_blank=True)
    dv_complexity.add('D2:D1048576') # Apply to column D (Complexity) from row 2 onwards
    sheet.add_data_validation(dv_complexity)

    # Create data validation for 'species_names'
    # if species_names_list:
    #     dv_species = DataValidation(type="list", formula1='"' + ','.join(species_names_list) + '"', allow_blank=True)
    #     dv_species.add('H2:H1048576') # Apply to column H (species_names) from row 2 onwards
    #     sheet.add_data_validation(dv_species)
    


    # Add comments to guide users for multi-select fields
    sheet['H1'].comment = Comment("For multiple species, separate names with commas (e.g., 'Species A, Species B')", "Admin")


    # Write data
    for skill in skills:
        species_names = ', '.join([s.name for s in skill.species])

        sheet.append([
            skill.name,
            skill.description,
            skill.validity_period_months,
            skill.complexity.value,
            skill.reference_urls_text,
            skill.training_videos_urls_text,
            skill.potential_external_tutors_text,
            species_names
        ])
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'skills_export_{datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")}.xlsx')



@bp.route('/training_sessions/create', methods=['GET', 'POST'])
@login_required
@permission_required('training_session_manage')
def create_training_session():
    form = TrainingSessionForm()

    # Pre-population from training requests (single or multiple)
    request_ids_str = request.args.get('request_ids')
    species_id_param = request.args.get('species_id')
    user_ids_param = request.args.get('user_ids')
    skill_ids_param = request.args.get('skill_ids')

    prefill_users = []
    prefill_skills = []
    prefill_species = None

    if request.method == 'GET':
        if request_ids_str:
            request_ids = [int(rid) for rid in request_ids_str.split(',') if rid.isdigit()]
            training_requests = TrainingRequest.query.filter(TrainingRequest.id.in_(request_ids)).all()

            unique_users = set()
            unique_skills = set()
            unique_species = set()

            for req in training_requests:
                unique_users.add(req.requester)
                for skill in req.skills_requested:
                    unique_skills.add(skill)
                for species in req.species_requested:
                    unique_species.add(species)

            prefill_users = list(unique_users)
            prefill_skills = list(unique_skills)

            if len(unique_species) == 1:
                prefill_species = unique_species.pop()
            elif len(unique_species) > 1:
                flash('Multiple species requested across selected training requests. Please select the main species manually.', 'warning')

            form.attendees.data = prefill_users
            form.skills_covered.data = prefill_skills
            if prefill_species:
                form.main_species.data = prefill_species

        elif species_id_param or user_ids_param or skill_ids_param:
            if species_id_param:
                species_obj = Species.query.get(int(species_id_param))
                if species_obj:
                    prefill_species = species_obj
                    form.main_species.data = species_obj
            
            if user_ids_param:
                user_ids_list = [int(uid) for uid in user_ids_param.split(',') if uid.isdigit()]
                users_obj = User.query.filter(User.id.in_(user_ids_list)).all()
                if users_obj:
                    prefill_users = users_obj
                    form.attendees.data = users_obj
            
            if skill_ids_param:
                skill_ids_list = [int(sid) for sid in skill_ids_param.split(',') if sid.isdigit()]
                skills_obj = Skill.query.filter(Skill.id.in_(skill_ids_list)).all()
                if skills_obj:
                    prefill_skills = skills_obj
                    form.skills_covered.data = skills_obj

    if form.validate_on_submit():
        session = TrainingSession(
            title=form.title.data,
            location=form.location.data,
            start_time=form.start_time.data,
            end_time=form.end_time.data,
            main_species=Species.query.get(request.form.get('main_species')),
            ethical_authorization_id=form.ethical_authorization_id.data,
            animal_count=form.animal_count.data
        )
        db.session.add(session)
        
        # Process dynamic skill-tutor rows
        skill_ids = set()
        tutor_ids = set()
        tutor_skill_pairs = []

        for key in request.form:
            if key.startswith('program-') and key.endswith('-skill'):
                row_index = key.split('-')[1]
                skill_id = request.form.get(key)
                tutor_id = request.form.get(f'program-{row_index}-tutor')
                if skill_id and tutor_id:
                    skill_ids.add(int(skill_id))
                    tutor_ids.add(int(tutor_id))
                    tutor_skill_pairs.append({'skill_id': int(skill_id), 'tutor_id': int(tutor_id)})

        if skill_ids:
            session.skills_covered = Skill.query.filter(Skill.id.in_(skill_ids)).all()
        
        if tutor_ids:
            session.tutors = User.query.filter(User.id.in_(tutor_ids)).all()

        db.session.add(session)
        db.session.flush() # Flush to get session.id

        # Handle tutor-skill mapping
        for pair in tutor_skill_pairs:
            tutor_skill = TrainingSessionTutorSkill(
                training_session_id=session.id,
                tutor_id=pair['tutor_id'],
                skill_id=pair['skill_id']
            )
            db.session.add(tutor_skill)

        if form.attachment.data:
            filename = secure_filename(form.attachment.data.filename)
            upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'training_sessions')
            os.makedirs(upload_folder, exist_ok=True)
            file_path = os.path.join(upload_folder, filename)
            form.attachment.data.save(file_path)
            session.attachment_path = os.path.join('uploads', 'training_sessions', filename)

        session.attendees = form.attendees.data
        
        db.session.commit()

        flash('Session de formation créée avec succès !', 'success')
        return redirect(url_for('admin.manage_training_sessions'))
    
    # Update JSON prefill variables for the template
    prefill_users_json = json.dumps([u.id for u in prefill_users])
    prefill_skills_json = json.dumps([s.id for s in prefill_skills])
    prefill_species_id = prefill_species.id if prefill_species else "null"

    return render_template('admin/create_training_session.html', title='Create Training Session', form=form, session=None,
                           prefill_users_json=prefill_users_json,
                           prefill_skills_json=prefill_skills_json,
                           prefill_species_id=prefill_species_id,
                           api_key=current_user.api_key)



@bp.route('/training_sessions')
@login_required
@permission_required('training_session_manage')
def manage_training_sessions():
    filter_param = request.args.get('filter')
    query = TrainingSession.query

    if filter_param == 'to_be_finalized':
        query = query.filter(
            TrainingSession.start_time < datetime.now(timezone.utc),
            TrainingSession.status != 'Realized'
        )

    training_sessions = query.order_by(TrainingSession.start_time.desc()).all()
    return render_template('admin/manage_training_sessions.html', title='Manage Training Sessions', training_sessions=training_sessions)

@bp.route('/training_sessions/edit/<int:session_id>', methods=['GET', 'POST'])
@login_required
@permission_required('training_session_manage')
def edit_training_session(session_id):
    session = TrainingSession.query.get_or_404(session_id)
    form = TrainingSessionForm(obj=session) # Pre-populate form with session data
    if form.validate_on_submit():
        session.title = form.title.data
        session.location = form.location.data
        session.start_time = form.start_time.data
        session.end_time = form.end_time.data
        session.main_species = Species.query.get(request.form.get('main_species'))
        session.ethical_authorization_id = form.ethical_authorization_id.data
        session.animal_count = form.animal_count.data
        db.session.add(session)
        
        # Process dynamic skill-tutor rows
        skill_ids = set()
        tutor_ids = set()
        tutor_skill_pairs = []

        for key in request.form:
            if key.startswith('program-') and key.endswith('-skill'):
                row_index = key.split('-')[1]
                skill_id = request.form.get(key)
                tutor_id = request.form.get(f'program-{row_index}-tutor')
                if skill_id and tutor_id:
                    skill_ids.add(int(skill_id))
                    tutor_ids.add(int(tutor_id))
                    tutor_skill_pairs.append({'skill_id': int(skill_id), 'tutor_id': int(tutor_id)})

        if skill_ids:
            session.skills_covered = Skill.query.filter(Skill.id.in_(skill_ids)).all()
        
        if tutor_ids:
            session.tutors = User.query.filter(User.id.in_(tutor_ids)).all()

        db.session.add(session)
        db.session.flush() # Flush to get session.id

        # Handle tutor-skill mapping
        for pair in tutor_skill_pairs:
            tutor_skill = TrainingSessionTutorSkill(
                training_session_id=session.id,
                tutor_id=pair['tutor_id'],
                skill_id=pair['skill_id']
            )
            db.session.add(tutor_skill)

        if form.attachment.data:
            filename = secure_filename(form.attachment.data.filename)
            upload_folder = os.path.join(current_app.root_path, 'static', 'uploads', 'training_sessions')
            os.makedirs(upload_folder, exist_ok=True)
            file_path = os.path.join(upload_folder, filename)
            form.attachment.data.save(file_path)
            session.attachment_path = os.path.join('uploads', 'training_sessions', filename)

        session.attendees = form.attendees.data
        
        db.session.commit()

        flash('Session de formation mise à jour avec succès !', 'success')
        return redirect(url_for('admin.manage_training_sessions'))
    
    return render_template('admin/create_training_session.html', title='Edit Training Session', form=form, session=session)

@bp.route('/training_sessions/delete/<int:session_id>', methods=['POST'])
@login_required
@permission_required('training_session_manage')
def delete_training_session(session_id):
    session = TrainingSession.query.get_or_404(session_id)
    db.session.delete(session)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Session de formation supprimée avec succès.'})

@bp.route('/training_sessions/<int:session_id>', methods=['GET'])
@login_required
@permission_required('training_session_manage') # Assuming only admins can view all session details
def view_training_session_details(session_id):
    session = TrainingSession.query.options(
        db.joinedload(TrainingSession.attendees),
        db.joinedload(TrainingSession.skills_covered),
        db.joinedload(TrainingSession.main_species),
        db.joinedload(TrainingSession.tutors),
        db.joinedload(TrainingSession.tutor_skill_mappings)
    ).get_or_404(session_id)

    read_only = False
    if not current_user.can('training_session_manage'): # Check for specific permission
        if current_user not in session.attendees and current_user not in session.tutors:
            abort(403) # Not authorized
        read_only = True # Attendee or tutor, but not admin, so read-only

    return render_template('admin/view_training_session_details.html', 
                           title='Session Details', 
                           session=session,
                           read_only=read_only)

@bp.route('/training_sessions/<int:session_id>/validate', methods=['GET', 'POST'])
@login_required
@permission_required('training_session_validate')
def validate_training_session(session_id):
    session = TrainingSession.query.options(
        db.joinedload(TrainingSession.attendees),
        db.joinedload(TrainingSession.skills_covered)
    ).get_or_404(session_id)

    # is_admin = current_user.is_admin # No longer needed, use can()
    is_session_tutor = current_user in session.tutors
    can_validate_session = current_user.can('training_session_validate') or is_session_tutor

    authorized_skills_for_current_user = set()
    if current_user.can('training_session_manage'): # Admins can validate all skills
        authorized_skills_for_current_user.update([skill.id for skill in session.skills_covered])
    elif is_session_tutor:
        for mapping in session.tutor_skill_mappings:
            if mapping.tutor_id == current_user.id:
                authorized_skills_for_current_user.add(mapping.skill_id)

    if request.method == 'POST':
        if not can_validate_session:
            flash('You are not authorized to validate competencies for this session.', 'danger')
            return redirect(url_for('admin.validate_training_session', session_id=session.id))

        # Manually parse form data
        for attendee in session.attendees:
            for skill in session.skills_covered:
                # Check if the current user is authorized to validate this skill
                if not current_user.can('training_session_manage') and skill.id not in authorized_skills_for_current_user:
                    continue

                # Check if this competency was submitted
                acquired_field_name = f'acquired-{attendee.id}-{skill.id}'
                if acquired_field_name in request.form:
                    level_field_name = f'level-{attendee.id}-{skill.id}'
                    level = request.form.get(level_field_name)

                    # Get species associated with the skill from the training session
                    # Assuming skill.species gives the species for this skill
                    skill_species = skill.species # This is a list of Species objects
                    skill_species_ids = sorted([s.id for s in skill_species])

                    # Try to find an existing competency for this user and skill
                    existing_competencies = Competency.query.filter(
                        Competency.user_id == attendee.id,
                        Competency.skill_id == skill.id
                    ).all()

                    competency_to_update = None
                    for comp in existing_competencies:
                        comp_species_ids = sorted([s.id for s in comp.species])
                        if comp_species_ids == skill_species_ids:
                            competency_to_update = comp
                            break

                    if competency_to_update:
                        # Update existing competency
                        competency_to_update.level = level
                        competency_to_update.evaluation_date = datetime.now(timezone.utc)
                        competency_to_update.evaluator = current_user
                        competency_to_update.training_session = session
                        competency_to_update.external_evaluator_name = None # Ensure this is cleared if an internal evaluator is used
                        competency_to_update.external_training_id = None # Ensure this is cleared for training sessions
                        db.session.add(competency_to_update)
                    else:
                        # Create new competency
                        competency = Competency(
                            user=attendee,
                            skill=skill,
                            level=level,
                            evaluation_date=datetime.now(timezone.utc),
                            evaluator=current_user,
                            training_session=session,
                            external_training_id=None # Ensure this is None for training sessions
                        )
                        db.session.add(competency)
                        db.session.flush() # Flush to assign an ID to the new competency before modifying its relationships
                        competency.species = skill_species # Associate species with the new competency

        db.session.commit()

        # Check if session is fully validated
        all_skills_validated = True
        for attendee in session.attendees:
            for skill in session.skills_covered:
                competency = Competency.query.filter_by(
                    user_id=attendee.id,
                    skill_id=skill.id,
                    training_session_id=session.id
                ).first()
                if not competency or not competency.evaluation_date:
                    all_skills_validated = False
                    break
            if not all_skills_validated:
                break

        if all_skills_validated:
            session.status = 'Realized'
            db.session.add(session)
            db.session.commit()
            flash('Session de formation entièrement validée et réalisée !', 'success')
        else:
            flash('Compétences validées avec succès (session non entièrement réalisée).', 'success')

        return redirect(url_for('admin.validate_training_session', session_id=session.id))

    # GET request
    # Prepare data for the template
    attendees_data = []
    for attendee in session.attendees:
        skills_data = []
        for skill in session.skills_covered:
            competency = Competency.query.filter_by(
                user_id=attendee.id,
                skill_id=skill.id,
                training_session_id=session.id
            ).first()
            skills_data.append({
                'skill': skill,
                'competency': competency
            })
        attendees_data.append({
            'attendee': attendee,
            'skills': skills_data
        })

    return render_template('admin/validate_training_session.html',
                           title='Validate Training Session',
                           session=session,
                           attendees_data=attendees_data,
                           can_validate_session=can_validate_session,
                           authorized_skills_for_current_user=authorized_skills_for_current_user,
                           is_admin=current_user.can('admin_access')) # Pass admin_access permission

@bp.route('/training_requests/reject/<int:request_id>', methods=['POST'])
@login_required
@permission_required('training_request_manage')
def reject_training_request(request_id):
    training_request = TrainingRequest.query.get_or_404(request_id)
    training_request.status = TrainingRequestStatus.REJECTED
    db.session.commit()
    return jsonify({'success': True, 'message': 'Demande de formation rejetée avec succès.'})

# External Training Validation
@bp.route('/validate_external_trainings')
@login_required
@permission_required('external_training_validate')
def validate_external_trainings():
    pending_external_trainings = ExternalTraining.query.filter_by(status=ExternalTrainingStatus.PENDING).all()
    return render_template('admin/validate_external_trainings.html', title='Validate External Trainings', trainings=pending_external_trainings)

@bp.route('/validate_external_trainings/approve/<int:training_id>', methods=['POST'])
@login_required
@permission_required('external_training_validate')
def approve_external_training(training_id):
    external_training = ExternalTraining.query.options(db.joinedload(ExternalTraining.skill_claims).joinedload(ExternalTrainingSkillClaim.skill)).get_or_404(training_id)
    external_training.status = ExternalTrainingStatus.APPROVED
    external_training.validator = current_user
    db.session.add(external_training)

    # Create or update competencies for each skill claim
    for skill_claim in external_training.skill_claims:
        # Check for existing competency for this user, skill, and species combination
        existing_competency = Competency.query.filter(
            Competency.user_id == external_training.user.id,
            Competency.skill_id == skill_claim.skill.id
        ).first()

        # Filter by species. This is a bit complex due to many-to-many relationship.
        # We need to check if the existing competency has the exact same set of species as the skill_claim.
        # For simplicity, we'll first check if a competency exists for the user and skill.
        # If it does, we'll update it. If the species are different, we'll treat it as a new competency.
        # This approach might lead to multiple competencies for the same skill if species differ, which aligns with the user's request.

        competency_found = False
        if existing_competency:
            # Check if the species associated with the existing competency are the same as the skill claim
            existing_species_ids = sorted([s.id for s in existing_competency.species])
            claim_species_ids = sorted([s.id for s in skill_claim.species_claimed])

            if existing_species_ids == claim_species_ids:
                # Update existing competency
                existing_competency.level = skill_claim.level
                existing_competency.evaluation_date = datetime.now(timezone.utc)
                if external_training.external_trainer_name:
                    existing_competency.external_evaluator_name = external_training.external_trainer_name
                    existing_competency.evaluator = None
                else:
                    existing_competency.evaluator = current_user
                    existing_competency.external_evaluator_name = None
                db.session.add(existing_competency)
                competency_found = True

        if not competency_found:
            # Create new competency
            competency = Competency(
                user=external_training.user,
                skill=skill_claim.skill,
                level=skill_claim.level,
                evaluation_date=external_training.date,
                evaluator=None,
                external_evaluator_name=None
            )

            if external_training.external_trainer_name:
                competency.external_evaluator_name = external_training.external_trainer_name
                competency.evaluator = None
            else:
                competency.evaluator = current_user
                competency.external_evaluator_name = None

            db.session.add(competency)
            db.session.flush()

            competency.external_training_id = external_training.id # Set external_training_id

            # Transfer species_claimed from ExternalTrainingSkillClaim to Competency
            competency.species = skill_claim.species_claimed

        # If user wants to be a tutor, add them to the skill's tutors
        if skill_claim.wants_to_be_tutor and external_training.user not in skill_claim.skill.tutors:
            skill_claim.skill.tutors.append(external_training.user)
        
        # If a practice date is provided, create a skill practice event
        if skill_claim.practice_date:
            practice_event = SkillPracticeEvent(
                user=external_training.user,
                practice_date=skill_claim.practice_date,
                notes="Practice declared from external training validation."
            )
            practice_event.skills.append(skill_claim.skill)
            db.session.add(practice_event)
    
    db.session.commit()
    flash('External training approved and competencies created!', 'success')
    return redirect(url_for('admin.validate_external_trainings'))

@bp.route('/validate_external_trainings/reject/<int:training_id>', methods=['POST'])
@login_required
@permission_required('external_training_validate')
def reject_external_training(training_id):
    external_training = ExternalTraining.query.get_or_404(training_id)
    external_training.status = ExternalTrainingStatus.REJECTED
    external_training.validator = current_user
    db.session.add(external_training)
    db.session.commit()
    flash('External training rejected.', 'info')
    return redirect(url_for('admin.validate_external_trainings'))

from collections import defaultdict

@bp.route('/training_requests')
@login_required
@permission_required('training_request_manage')
def list_training_requests():
    pending_training_requests = TrainingRequest.query.options(
        db.joinedload(TrainingRequest.requester),
        db.joinedload(TrainingRequest.skills_requested).joinedload(Skill.species),
        db.joinedload(TrainingRequest.species_requested)
    ).filter_by(status=TrainingRequestStatus.PENDING).all()

    requests_by_species = defaultdict(list)
    for req in pending_training_requests:
        # A request can be associated with multiple species through its skills.
        # We add the request to each associated species group.
        associated_species = req.associated_species
        if not associated_species:
            # If no species is associated (e.g. skill has no species), group it under a special key
            if None not in requests_by_species:
                class MockSpecies:
                    id = 0
                    name = "Sans Espèce Spécifiée"
                requests_by_species[MockSpecies()] = []
            requests_by_species[next(s for s in requests_by_species if s.id == 0)].append(req)
        else:
            for species in associated_species:
                # To avoid using the species object itself as a key, which can be tricky,
                # we can use species.id and pass a dictionary of species objects separately.
                requests_by_species[species].append(req)

    return render_template('admin/list_training_requests.html',
                           title='Pending Training Requests',
                           requests_by_species=requests_by_species)

# Reports (Placeholder)
@bp.route('/tutor_less_skills_report')
@login_required
@permission_required('view_reports')
def tutor_less_skills_report():
    # This will require a more complex query to find skills with no associated tutors
    skills_without_tutors = Skill.query.filter(~Skill.tutors.any()).all()
    return render_template('admin/tutor_less_skills_report.html', title='Skills Without Tutors Report', skills=skills_without_tutors)


@bp.route('/recycling_report')
@login_required
@permission_required('view_reports')
def recycling_report():
    report_data = defaultdict(lambda: defaultdict(list))
    all_users = User.query.options(db.joinedload(User.competencies).joinedload(Competency.skill).joinedload(Skill.species)).all()

    for user in all_users:
        for comp in user.competencies:
            if comp.needs_recycling:
                # This competency is expired.
                # A competency can be for multiple species.
                if comp.species:
                    for species in comp.species:
                        if user not in report_data[species][comp.skill]:
                            report_data[species][comp.skill].append(user)
                else:
                    # If competency has no species, check the skill's species
                    if comp.skill.species:
                        for species in comp.skill.species:
                            if user not in report_data[species][comp.skill]:
                                report_data[species][comp.skill].append(user)
                    else:
                        if None not in report_data:
                            class MockSpecies:
                                id = 0
                                name = "Sans Espèce Spécifiée"
                            report_data[None] = defaultdict(list)
                        if user not in report_data[None][comp.skill]:
                            report_data[None][comp.skill].append(user)

    # Create a mock species for the 'None' key if it exists
    if None in report_data:
        class MockSpecies:
            id = 0
            name = "Sans Espèce Spécifiée"
        report_data[MockSpecies()] = report_data.pop(None)

    return render_template('admin/recycling_report.html', title='Rapport de Recyclage', report_data=report_data)

@bp.route('/continuous_training_compliance_report')
@login_required
@permission_required('view_reports')
def continuous_training_compliance_report():
    users = User.query.all()
    report_data = []
    for user in users:
        report_data.append({
            'user': user,
            'total_hours': user.total_continuous_training_hours_6_years,
            'live_hours': user.live_continuous_training_hours_6_years,
            'online_hours': user.online_continuous_training_hours_6_years,
            'required_hours': user.required_continuous_training_hours,
            'is_compliant': user.is_continuous_training_compliant,
            'live_ratio': user.live_training_ratio,
            'is_live_ratio_compliant': user.is_live_training_ratio_compliant,
            'is_at_risk_next_year': user.is_at_risk_next_year
        })
    return render_template('admin/continuous_training_compliance_report.html', title='Rapport de Conformité Formation Continue', report_data=report_data)

@bp.route('/proposed_skills')
@login_required
@permission_required('skill_manage') # Assuming proposed skills are managed by skill managers
def proposed_skills():
    proposed = TrainingRequest.query.filter_by(status=TrainingRequestStatus.PROPOSED_SKILL).order_by(TrainingRequest.request_date.desc()).all()
    return render_template('admin/proposed_skills.html', title='Proposed Skills', proposed_skills=proposed)

@bp.route('/api/training_path/<int:path_id>/skills')
@login_required
@permission_required('training_path_manage')
def get_training_path_skills(path_id):
    training_path = TrainingPath.query.options(db.joinedload(TrainingPath.skills_association).joinedload(TrainingPathSkill.skill)).get_or_404(path_id)
    skills_data = []
    for tps in training_path.skills_association:
        skills_data.append({
            'id': tps.skill.id,
            'name': tps.skill.name,
            'description': tps.skill.description,
            'species': training_path.species.name # Use the species from TrainingPath
        })
    return jsonify(skills_data)
